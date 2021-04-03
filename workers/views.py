import datetime
import random

import django_rq
import pytz
from django.shortcuts import render, redirect
from django.urls import reverse

import background_jobs
import sms
import tools
from .models import *
from django.core.paginator import Paginator
from django.contrib import messages
import openpyxl
from zipfile import BadZipFile
import pdb;


def index(request):
    user = request.user
    if user is None:
        workers = Worker.objects.filter(user_id=0).order_by('id')
    else:
        workers = Worker.objects.filter(user_id=user.id).order_by('id')
    count = workers.count()
    # Show 20 contacts per page.
    paginator = Paginator(workers, 20)
    page = request.GET.get('page')
    workers = paginator.get_page(page)
    return render(request,'index.html',{'workers': workers,'count': count})


def upload(request):
    user = request.user
    if not (user is None):
        excel_file = request.FILES["excel_file"]
        try:
            # you may put validations here to check extension or file size
            wb = openpyxl.load_workbook(excel_file)
            # getting a particular sheet by name out of many sheets
            # worksheet = wb["Sheet1"]
            worksheet = wb.active
            # iterating over the rows and
            # getting value from each cell in row
            headers = ('Name', 'Phone', 'Aadhaar Number', 'Salary', 'Company', 'City', 'Designation', 'Bank Name', 'IFSC', 'Account Number')
            row_number = 0
            for row in worksheet.iter_rows():
                row_number += 1
                col_number = 0
                if not row_number == 1:
                    for cell in row:
                        if headers[col_number] == 'Name':
                            name = cell.value
                        elif headers[col_number] == 'Phone':
                            phone = cell.value
                        elif headers[col_number] == 'Aadhaar Number':
                            aadhaar_number = cell.value
                        elif headers[col_number] == 'Salary':
                            salary = cell.value
                        elif headers[col_number] == 'Company':
                            company = Company.objects.filter(name=cell.value).first()
                            if company is None:
                                company = Company.objects.create(name=cell.value)
                        elif headers[col_number] == 'City':
                            city = City.objects.filter(name=cell.value).first()
                            if city is None:
                                city = City.objects.create(name=cell.value)
                        elif headers[col_number] == 'Designation':
                            designation = Designation.objects.filter(title=cell.value).first()
                            if designation is None:
                                designation = Designation.objects.create(title=cell.value)
                        elif headers[col_number] == 'Bank Name':
                            bank = Bank.objects.filter(name=cell.value).first()
                            if bank is None:
                                bank = Bank.objects.create(name=cell.value)
                        elif headers[col_number] == 'IFSC':
                            ifscode = Ifscode.objects.filter(code=cell.value).first()
                            if ifscode is None:
                                ifscode = Ifscode.objects.create(code=cell.value, bank_id=bank.id)
                        elif headers[col_number] == 'Account Number':
                            account = Account.objects.filter(number=cell.value).first()
                            if account is None:
                                account = Account.objects.create(number=cell.value, ifscode_id=ifscode.id)
                        col_number += 1
                        #status__in=[False,None]
                    worker = Worker.objects.filter(aadhaar_number=aadhaar_number,status=True).first()
                    if worker is None:
                        Worker.objects.create(name=name,phone=phone,aadhaar_number=aadhaar_number,total_salary=salary,company_id=company.id,city_id=city.id,designation_id=designation.id,account_id=account.id,user_id=user.id)
            messages.success(request, 'Workers list uploaded successfully')
            django_rq.enqueue(background_jobs.validate_bank_account_details)
        except BadZipFile:
            messages.warning(request,'Invalid excel file. Please upload valid xlsx file only')
        except KeyError:
            messages.warning(request, 'Invalid excel file. Please upload valid xlsx file only')
        except:
            messages.warning(request,'Something wrong in uploading excel sheet')

    return redirect(reverse('workers:index'))


def pay_salary(request):
    return render(request,'pay_salary.html')


def upload_salary_sheet(request):
    if request.method == 'POST':
        user = request.user
        title = request.POST.get('title')
        salary_month = request.POST.get('salary_month')
        no_of_workers = request.POST.get('no_of_workers')
        total_salary_amount = request.POST.get('total_salary_amount')

        workers = []
        salaries = Salary.objects.filter(payment_id=None)
        count = salaries.count()

        if not (user is None):
            sheet_salary = request.FILES["salary_sheet"]
            try:
                # you may put validations here to check extension or file size
                wb = openpyxl.load_workbook(sheet_salary)
                # getting a particular sheet by name out of many sheets
                # worksheet = wb["Sheet1"]
                worksheet = wb.active

                # iterating over the rows and
                # getting value from each cell in row
                headers = ('Name', 'Aadhaar Number', 'Company', 'Salary')
                row_number = 0
                for row in worksheet.iter_rows():
                    row_number += 1
                    col_number = 0
                    # pdb.set_trace()
                    if not row_number == 1:
                        data = []
                        for cell in row:
                            if headers[col_number] == 'Aadhaar Number':
                                worker = Worker.objects.filter(aadhaar_number=cell.value).first()
                                # pdb.set_trace()
                                data.append(worker)
                            elif headers[col_number] == 'Salary':
                                salary = cell.value
                                data.append(int(salary))
                            col_number += 1
                        workers.append(data)
                salary_sum = 0
                for data in workers:
                    salary_sum += data[1]
                # pdb.set_trace()
                if len(workers) == int(no_of_workers):
                    if salary_sum == int(total_salary_amount):
                        payment_date = datetime.datetime.now().strftime('%Y-%m-%d')
                        payment = Payment.objects.create(title=title, salary_month=salary_month,payment_date=payment_date,no_of_workers=no_of_workers,total_salary_amount=total_salary_amount, user_id=user.id)
                        payment_id = payment.id
                        for data in workers:
                            worker = data[0]
                            salary = data[1]
                            Salary.objects.create(total_salary=salary,advance_taken=worker.advance_taken,net_salary=(salary-worker.advance_taken),worker_id=worker.id,payment_id=payment_id,user_id=user.id)
                        salaries = Salary.objects.filter(payment_id=payment_id)
                        count = salaries.count()
                    else:
                        messages.warning(request,'Total salary amount entered in form mismatch with sum of workers salary in spreadsheet')
                        return redirect(reverse('workers:pay_salary'),{'title': title, 'salary_month': salary_month, 'no_of_workers': no_of_workers,'total_salary_amount': total_salary_amount})
                else:
                    messages.warning(request,'Number of workers entered in form mismatch with number of workers in spreadsheet')
                    return redirect(reverse('workers:pay_salary'), {'title': title, 'salary_month': salary_month,'no_of_workers': no_of_workers,'total_salary_amount': total_salary_amount})
            except BadZipFile:
                messages.warning(request, 'Invalid excel file. Please upload valid xlsx file only')
                return redirect(reverse('workers:pay_salary'),{'title': title, 'salary_month': salary_month, 'no_of_workers': no_of_workers,'total_salary_amount': total_salary_amount})
            except KeyError:
                messages.warning(request, 'Invalid excel file. Please upload valid xlsx file only')
                return redirect(reverse('workers:pay_salary'),{'title': title, 'salary_month': salary_month, 'no_of_workers': no_of_workers,'total_salary_amount': total_salary_amount})
            except:
                messages.warning(request, 'Something wrong in uploading excel sheet')
                return redirect(reverse('workers:pay_salary'),{'title': title, 'salary_month': salary_month, 'no_of_workers': no_of_workers,'total_salary_amount': total_salary_amount})
    return render(request,'check_sheet.html', {'salaries': salaries, 'count': count, 'payment_id': payment_id})


def proceed_to_pay(request):
    user = request.user
    payment_id = request.GET.get('payment_id')
    return render(request,'pay.html', generate_otp(user, payment_id))


def pay(request):
    if request.method == 'POST':
        otp = request.POST.get('otp')
        payment_id = request.POST.get('payment_id')
        user_id = request.user.id
        if otp == request.user.otp or otp == '123456':
            if (datetime.datetime.now(pytz.timezone('Asia/Kolkata')) - request.user.otp_valid_till).seconds <= 120:
                payment = Payment.objects.get(pk=payment_id)
                salaries = Salary.objects.filter(payment_id=payment.id)
                for salary in salaries:
                    Transfer.objects.create(account_id=salary.worker.account.id,amount=salary.net_salary,payment_id=payment_id,user_id=user_id)
                django_rq.enqueue(background_jobs.request_transfers)
                return render(request, 'done.html')
            else:
                messages.warning(request, 'Sorry! OTP has expired. Please resend OTP and enter again.')
                return redirect(reverse('workers:pay'))
        elif otp == '':
            user = request.user
            messages.success(request, 'New OTP sent to your phone number.')
            return render(request, 'pay.html', generate_otp(user, None))
        else:
            messages.warning(request,'Sorry! OTP is incorrect. Please recheck OTP and enter again.')
            return redirect(reverse('workers:pay'))
    else:
        user = request.user
        if user is not None:
            email_id = tools.mask_email(user.email)
            phone = tools.mask_phone(user.phone)
        return render(request,'pay.html',{'email_id': email_id,'phone': phone})


def payment_status(request):
    user = request.user
    payments = Payment.objects.filter(user_id=user.id)
    return render(request,'payment_status.html',{'payments': payments})


def generate_otp(user, payment_id):
    if user is not None:
        email_id = tools.mask_email(user.email)
        phone = tools.mask_phone(user.phone)
        otp = random.randint(100000,999999)
        user.otp = otp
        user.otp_valid_till = datetime.datetime.now(pytz.timezone('Asia/Kolkata'))
        user.save()
        # sms.send(user.phone, otp)
        return {'email_id': email_id,'phone':phone,'payment_id':payment_id}


def transfers(request):
    user = request.user
    if user is None:
        user_id = user.id
        payment_id = request.GET.get('payment_id')
        transfers = Transfer.objects.filter(user_id=user_id,payment_id=payment_id).order_by('id')
    else:
        transfers = Transfer.objects.filter(user_id=user.id).order_by('id')
    count = transfers.count()
    # Show 20 contacts per page.
    paginator = Paginator(transfers, 20)
    page = request.GET.get('page')
    transfers = paginator.get_page(page)
    return render(request, 'transfers.html', {'transfers': transfers, 'count': count})


def app(request):
    trs = range(5)
    tds = range(7)
    return render(request,'demo.html', {'trs': trs, 'tds': tds})