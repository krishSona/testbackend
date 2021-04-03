from django.http import JsonResponse
from django.shortcuts import get_object_or_404

from rest_framework import viewsets, views
from django.contrib.postgres.search import SearchVector
import datetime

from api.v1.employee.validations import validate_employee_data

import utilities
from api.v1.authentication.views import get_tokens_for_user

from api.v1.employee.serializers import (
    EmployeeListSerializer,
    EmployeeCreateSerializer,
    EmployeeUpdateSerializer
)
from api.v1.services.views import get_utc_datetime

from core.models import (
    Employee,
    Ifs,
    Bank,
    Attendance,
)

from daily_salary import settings

from authentication.models import User
from drf_yasg2.utils import swagger_auto_schema
from drf_yasg2 import openapi
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError

from services.sms.send import call

import openpyxl
from zipfile import BadZipFile

import background_jobs


def employee_home_details(instance):
    serializer = EmployeeListSerializer(instance)
    data = serializer.data

    # to find working day
    today = datetime.datetime.now()
    day = today.strftime("%A")[:3]
    working_days = instance.work_days.get('days')
    if day in working_days:
        is_working = True
    else:
        is_working = False
    data["is_working"] = is_working

    # to check attendance_marked
    today_date = str(today.date())
    employee_id = data.get('id', None)
    attendance = Attendance.objects.filter(
        date=today_date, employee=employee_id).first()
    if attendance:
        attendance_marked = True
    else:
        attendance_marked = False
    data["attendance_marked"] = attendance_marked

    curr_date = datetime.datetime.now().date()
    last_date_of_month = background_jobs.calculate_last_date_of_month(curr_date)
    days_left_month_end = last_date_of_month.day - curr_date.day
    month_end_notification = None
    if days_left_month_end < 3:
        month_end_notification = str(days_left_month_end + 1) + " days left to withdraw money from Wallet"
    data["month_end_notification"] = month_end_notification
    data["available_balance"] = instance.get_available_balance()
    data["service_status"] = True
    return data


def create_employee_response(employee, refresh):
    response_data = {
        "employee": {
            "id": employee.id,
            "employee_id": employee.employee_id,
            "phone": employee.phone,
            "name": employee.name,
            "email": employee.email if employee.email else None,
            "joining_date": employee.joining_date if employee.joining_date else None,
            "salary_day": employee.salary_day if employee.salary_day else None,
            "net_monthly_salary": employee.net_monthly_salary,
            "company": {
                "id": employee.company.id if employee.company else None,
                "name": employee.company.name if employee.company else None
            }
        },
        "status": "True",
        "message": "Logged in successfully",
        "balance": float(employee.balance) if employee.balance else 0.0,
        "token": {
            "refresh": str(refresh),
            "access": str(refresh.access_token),
            "x-karza-key": str(settings.CORE_X_KARZA_KEY)
        }
    }
    return response_data


class EmployeeViewSet(viewsets.ModelViewSet):
    queryset = Employee.objects.filter(deleted_at=None).order_by('-id')
    serializer_class = EmployeeListSerializer
    #permission_classes = [permissions.IsAuthenticated]

    lookup_fields = ("rid", "pk")

    def get_object(self):
        queryset = self.get_queryset()
        queryset = self.filter_queryset(queryset)
        filter = {}
        for field in self.lookup_fields:
            if self.kwargs.get(field):
                filter[field] = self.kwargs[field]
        obj = get_object_or_404(queryset, **filter)
        self.check_object_permissions(self.request, obj)
        return obj

    def list(self, request, *args, **kwargs):
        queryset = Employee.objects.filter(deleted_at=None).order_by('-id')
        search_text = self.request.query_params.get('search', None)
        phone = self.request.query_params.get('phone', None)
        if phone:
            queryset = queryset.filter(phone=phone)
        if search_text:
            queryset = queryset.annotate(
                        search=SearchVector('employee_id', 'name', 'phone'),
                     ).filter(search=search_text)
        page = self.paginate_queryset(queryset)
        serializer = EmployeeListSerializer(page, many=True)
        response = self.get_paginated_response(serializer.data)
        return response

    def create(self, request, *args, **kwargs):
        company_obj = validate_employee_data(data=request.data)
        name = request.data.get('name', None)
        email = request.data.get('email', None)
        if name:
            name = name.capitalize()
        if email:
            email = email.lower()
        employee_data = {
            "user": {
                "username": request.data.get('phone', None)
            },
            "employee_id": request.data.get('employee_id', None),
            "phone": request.data.get('phone', None),
            "name": name,
            "email": email,
            "joining_date": request.data.get('joining_date', None),
            "salary_day": request.data.get('salary_day', None),
            "company": company_obj.id,
            "net_monthly_salary": request.data.get('net_monthly_salary', None),
            "salary_type": request.data.get('salary_type', None),
            "agreed_with_terms_and_conditions": request.data.get(
                'agreed_with_terms_and_conditions', False),
            "confirmed": request.data.get('confirmed', False),
            "work_days": request.data.get('work_days', None)
        }
        phone = request.data.get('phone', None)
        user_exists = User.objects.filter(username=phone).first()
        employee_exists = Employee.objects.filter(phone=phone, deleted_at=None).first()
        # create an employee if user exists.
        if user_exists and not employee_exists:
            employee = Employee.objects.create(
                user=user_exists,
                employee_id=request.data.get('employee_id', None),
                phone=request.data.get('phone', None),
                name=name,
                email=email,
                joining_date=request.data.get('joining_date', None),
                salary_day=request.data.get('salary_day', None),
                company=company_obj,
                net_monthly_salary=request.data.get(
                    'net_monthly_salary', None),
                salary_type=request.data.get('salary_type', None),
                agreed_with_terms_and_conditions=request.data.get(
                    'agreed_with_terms_and_conditions', False),
                confirmed=request.data.get('confirmed', False),
                work_days=request.data.get('work_days', None)
            )
            refresh = get_tokens_for_user(user=user_exists)
            response_data = create_employee_response(employee, refresh)
            return JsonResponse(response_data)
        # if both user and employee exists.
        if user_exists and employee_exists:
            refresh = get_tokens_for_user(user=user_exists)
            response_data = create_employee_response(employee_exists, refresh)
            return JsonResponse(response_data)

        # if both user and employee does not exists.
        serializer = EmployeeCreateSerializer(data=employee_data)
        if serializer.is_valid():
            serializer.save()

            employee_obj = Employee.objects.get(id=serializer.data.get('id'))

            user = User.objects.get(username=request.data.get('phone'))
            refresh = get_tokens_for_user(user=user)
            response_data = create_employee_response(employee_obj, refresh)
            return JsonResponse(response_data)
        raise ValidationError({"message": str(serializer.errors)})

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        emp_data = {}
        if self.request.query_params.get('source') == "app" \
                and self.request.query_params.get('fieldset') == "home":
            emp_data['employee'] = employee_home_details(instance)
        else:
            emp_data = employee_home_details(instance)
        return JsonResponse(emp_data)

    def partial_update(self, request, *args, **kwargs):
        instance = self.get_object()
        source = self.request.query_params.get('source', None)
        fieldset = self.request.query_params.get('fieldset', None)
        device_name = request.data.get('device_name', None)
        device_id = request.data.get('device_id', None)
        ip_address = request.data.get('ip_address', None)
        loan_id = request.data.get('loan_id', None)

        digital_time_stamp = None
        if instance.name and device_name and device_id and ip_address:
            if not instance.kyc or not instance.e_nach:
                raise ValidationError({"message": "Please complete KYC and eNACH first."})
            digital_time_stamp = utilities.generate_digital_time_stamp(
                name=instance.name,
                device_name=device_name,
                device_id=device_id,
                ip_address=ip_address,
            )

        # to verify email of employee by token
        if source == "web" and fieldset == "verify_email":
            mail_token = request.data.get('mail_token', None)
            attendance_days = request.data.get('attendance_days', None)
            if not mail_token:
                raise ValidationError({
                    "verified": "False",
                    "message": "Mail token is required"
                })
            if mail_token:
                mail_token = str(mail_token).upper()
                employee_obj = Employee.objects.filter(mail_token=mail_token, deleted_at=None).first()
                if not employee_obj:
                    raise ValidationError({
                        "verified": "False",
                        "message": "Mail token has been expired"
                    })

            if attendance_days:
                if int(attendance_days) > employee_obj.calculate_work_day(datetime.datetime.now().day):
                    raise ValidationError({
                        "verified": "False",
                        "message": "Invalid Attendance days"})

            employee_obj.mail_verified = True
            employee_obj.mail_token = None
            employee_obj.save()

            # check if attendance already exists
            attendance = Attendance.objects.filter(
                employee=employee_obj
            ).first()
            if attendance:
                return JsonResponse({
                    "status": "True",
                    "message": "Email Address verified successfully."
                })

            # create today's attendance of employee if attendance does not exists
            daily_salary = employee_obj.get_daily_salary()
            balance = daily_salary * employee_obj.calculate_work_day(datetime.datetime.now().day)
            date_string = datetime.datetime.now().strftime('%-d, %Y')
            month_name = datetime.datetime.now().strftime('%b')
            date_string_ = month_name + " 1 to " + date_string
            description = "Salary credited for " + str(date_string_)
            attendance = Attendance.objects.create(
                status="present",
                salary=float(balance),
                description=description,
                employee=employee_obj,
                company=employee_obj.company,
            )

            # credit balance of employee in wallet
            employee_obj.balance = balance
            employee_obj.credited = balance
            employee_obj.save()

            # update verified attendance
            if employee_obj.email and employee_obj.mail_enabled and employee_obj.mail_verified:
                attendance.verified_salary = float(balance)
                attendance.save()

            if (not employee_obj.email) and attendance_days and int(attendance_days) > 0:
                attendance.verified_salary = int(attendance_days) * employee_obj.get_daily_salary()
                attendance.save()
            # SMS disabled
            # if not employee_obj.email:
            #     try:
            #         call(
            #             employee_obj.phone,
            #             "play_store_link",
            #             str(employee_obj.name).split(' ')[0],
            #             "https://tinyurl.com/sq5ewi5c",
            #         )
            #     except Exception as e:
            #         print(e)

            check_wallet_balance_url = "dailysalary://start"
            return JsonResponse({
                "verified": "True",
                "balance": int(balance),
                "url": check_wallet_balance_url,
                "company_name": utilities.get_company_first_name(employee_obj),
                "message": "Email Address verified successfully."
            })

        ifsc_dict = request.data.get('ifs', None)
        ifs = None
        if ifsc_dict:
            if not ifsc_dict.get('id'):
                bank = Bank.objects.filter(name=ifsc_dict.get('bank').get('name')).first()
                if not bank:
                    bank = Bank.objects.create(name=ifsc_dict.get('bank').get('name').upper())
                ifs = Ifs.objects.create(bank=bank, code=ifsc_dict.get('code').upper())
            else:
                ifs = Ifs.objects.get(id=ifsc_dict.get('id'))

        name = request.data.get('name', instance.name)
        email = request.data.get('email', instance.email)
        if name:
            name = name.capitalize()
        if email:
            email = email.lower()
        updated_data = {
            "name": name,
            "phone": request.data.get('phone', instance.phone),
            "email": email,
            "net_monthly_salary": request.data.get('net_monthly_salary', instance.net_monthly_salary),
            "salary_type": request.data.get('salary_type', instance.salary_type),
            "work_timings": request.data.get('work_timings', instance.work_timings),
            "work_days": request.data.get('work_days', instance.work_days),
            "agreed_with_terms_and_conditions": request.data.get(
                'agreed_with_terms_and_conditions', instance.agreed_with_terms_and_conditions),
            "confirmed": request.data.get('confirmed', instance.confirmed),
            "employee_id": request.data.get('employee_id', instance.employee_id),
            "joining_date": request.data.get('joining_date', instance.joining_date),
            "permanent_address": request.data.get('permanent_address', instance.permanent_address),
            "permanent_city": request.data.get('permanent_city', instance.permanent_city),
            "permanent_state": request.data.get('permanent_state', instance.permanent_state),
            "permanent_pincode": request.data.get('permanent_pincode', instance.permanent_pincode),
            "current_address": request.data.get('current_address', instance.current_address),
            "current_city": request.data.get('current_city', instance.current_city),
            "current_state": request.data.get('current_state', instance.current_state),
            "current_pincode": request.data.get('current_pincode', instance.current_pincode),
            "service_status": request.data.get('service_status', instance.service_status),
            "kyc": request.data.get('kyc', instance.kyc),
            "e_nach": request.data.get('e_nach', instance.e_nach),
            "wish_listing": request.data.get('wish_listing', instance.wish_listing),
            "mail_enabled": request.data.get('mail_enabled', instance.mail_enabled),
            "credit_limit": request.data.get('credit_limit', instance.credit_limit),
            "bank_account_number": request.data.get('bank_account_number', instance.bank_account_number),
            "ifs": ifs.id if ifs else (instance.ifs.id if instance.ifs else None),
            "balance": request.data.get('balance', instance.balance),
            "credited": request.data.get('credited', instance.credited),
            "debited": request.data.get('debited', instance.debited),
            "deleted_at": request.data.get('deleted_at', instance.deleted_at),
            "digital_time_stamp": digital_time_stamp if digital_time_stamp else instance.digital_time_stamp
        }
        serializer = EmployeeUpdateSerializer(instance, data=updated_data, partial=True)
        if serializer.is_valid():
            serializer.save()
            emp_data = employee_home_details(instance)
            if source == "app" and fieldset == "bank_account_details":
                return JsonResponse({"employee": emp_data})
            else:
                return JsonResponse(emp_data)
        raise ValidationError({"message": str(serializer.errors)})


class EmployeeBulkCreateView(views.APIView):
    def post(self, request, format=None):
        fields = self.request.data.get('fields', None)

        # validate employee's data from employee_sheet
        if fields == 'validate':
            employee_sheet = request.FILES["employee_sheet"]
            try:
                wb = openpyxl.load_workbook(employee_sheet)
                worksheet = wb.active
                headers = ('Name*', 'Phone*', 'Email', 'Net Monthly Salary*', 'Bank Name*',
                           'IFSC Code*', 'Bank Account Number*')
                employees_data = []
                row_number = 0
                for row in worksheet.iter_rows():
                    row_number += 1
                    col_number = 0
                    if not row_number == 1:
                        data = {}
                        for cell in row:
                            if headers[col_number] == 'Name*':
                                data['name'] = str(cell.value)
                            elif headers[col_number] == 'Phone*':
                                data['phone'] = str(cell.value)
                            elif headers[col_number] == 'Email':
                                data['email'] = str(cell.value)
                            elif headers[col_number] == 'Net Monthly Salary*':
                                data['net_monthly_salary'] = int(cell.value)
                            elif headers[col_number] == 'Bank Name*':
                                data['bank_name'] = str(cell.value)
                            elif headers[col_number] == 'IFSC Code*':
                                data['ifsc_code'] = str(cell.value)
                            elif headers[col_number] == 'Bank Account Number*':
                                data['bank_account_number'] = str(cell.value)
                            col_number += 1
                        employees_data.append(data)
            except BadZipFile:
                raise ValidationError({
                    'message': 'Invalid excel file. Please upload valid xlsx file only'})
            except KeyError:
                raise ValidationError({
                    'message': 'Invalid excel file. Please upload valid xlsx file only'})
            except:
                import traceback
                tr = traceback.format_exc()
                print("traceback:", tr)
                raise ValidationError({
                    'message': 'Something wrong in uploading excel sheet'})

            # validate employee's data
            validated_employees_count = 0
            for employee in employees_data:
                name = employee.get('name', None)
                phone = employee.get('phone', None)
                net_monthly_salary = employee.get('net_monthly_salary', None)
                bank_account_number = employee.get('bank_account_number', None)
                ifsc_code = employee.get('ifsc_code', None)
                bank_name = employee.get('bank_name', None)
                if not phone:
                    raise ValidationError({
                        "message": "Phone is required field",
                        "validated_employees_count": validated_employees_count
                    })
                if not name:
                    raise ValidationError({
                        "message": "Name is required field",
                        "validated_employees_count": validated_employees_count
                    })
                if not net_monthly_salary:
                    raise ValidationError({
                        "message": "Net Monthly Salary is required field",
                        "validated_employees_count": validated_employees_count
                    })
                if not bank_account_number:
                    raise ValidationError({
                        "message": "Bank Account Number is required field",
                        "validated_employees_count": validated_employees_count
                    })
                if not ifsc_code:
                    raise ValidationError({
                        "message": "IFSC code is required field",
                        "validated_employees_count": validated_employees_count
                    })
                if not bank_name:
                    raise ValidationError({
                        "message": "Bank Name is required field",
                        "validated_employees_count": validated_employees_count
                    })
                if phone:
                    if len(phone) != 10:
                        raise ValidationError({
                            "message": "Phone is not valid",
                            "validated_employees_count": validated_employees_count
                        })
                    users = User.objects.filter(username=phone)
                    if len(users) > 0:
                        raise ValidationError({
                            "message": "User already exist with this Phone.",
                            "validated_employees_count": validated_employees_count
                        })
                if name:
                    if len(name) > 50:
                        raise ValidationError({
                            "message": "Name is too long",
                            "validated_employees_count": validated_employees_count
                        })
                if ifsc_code:
                    if len(ifsc_code) != 11:
                        raise ValidationError({
                            "message": "IFSC code is not valid",
                            "validated_employees_count": validated_employees_count
                        })
                if bank_account_number:
                    if len(bank_account_number) < 9 or len(bank_account_number) > 18:
                        raise ValidationError({
                            "message": "Bank Account Number is not valid",
                            "validated_employees_count": validated_employees_count
                        })
                validated_employees_count += 1
            return JsonResponse({
                "message": "Employee Sheet Validated Successfully.",
                "validated_employees_count": validated_employees_count
            })

        # upload employee's data from employee_sheet
        elif fields == 'upload':
            employees_data = []
            employee_sheet = request.FILES["employee_sheet"]
            try:
                wb = openpyxl.load_workbook(employee_sheet)
                worksheet = wb.active
                headers = ('Name*', 'Phone*', 'Email', 'Net Monthly Salary*', 'Bank Name*',
                           'IFSC Code*', 'Bank Account Number*')
                row_number = 0
                for row in worksheet.iter_rows():
                    row_number += 1
                    col_number = 0
                    if not row_number == 1:
                        data = {}
                        for cell in row:
                            if headers[col_number] == 'Name*':
                                data['name'] = str(cell.value)
                            elif headers[col_number] == 'Phone*':
                                data['phone'] = str(cell.value)
                            elif headers[col_number] == 'Email':
                                data['email'] = str(cell.value)
                            elif headers[col_number] == 'Net Monthly Salary*':
                                data['net_monthly_salary'] = int(cell.value)
                            elif headers[col_number] == 'Bank Name*':
                                data['bank_name'] = str(cell.value)
                            elif headers[col_number] == 'IFSC Code*':
                                data['ifsc_code'] = str(cell.value)
                            elif headers[col_number] == 'Bank Account Number*':
                                data['bank_account_number'] = str(cell.value)
                            col_number += 1
                        employees_data.append(data)
            except BadZipFile:
                raise ValidationError({
                    'message': 'Invalid excel file. Please upload valid xlsx file only'})
            except KeyError:
                raise ValidationError({
                    'message': 'Invalid excel file. Please upload valid xlsx file only'})
            except:
                import traceback
                tr = traceback.format_exc()
                print("traceback:", tr)
                raise ValidationError({
                    'message': 'Something wrong in uploading excel sheet'})

            created_employees = []
            for employee in employees_data:
                username = employee.get('phone', None)
                ifsc_code = employee.get('ifsc_code', None)
                bank_name = employee.get('bank_name', None)

                user_data = {
                    "username": username
                }
                employee['user'] = user_data

                if bank_name:
                    bank = Bank.objects.filter(name=bank_name).first()
                    if bank is None:
                        bank = Bank.objects.create(name=bank_name)
                if bank_name and ifsc_code:
                    ifs = Ifs.objects.filter(code=ifsc_code).first()
                    if ifs is None:
                        ifs = Ifs.objects.create(code=ifsc_code, bank_id=bank.id)
                    employee['ifs'] = ifs.id

                # put employer's company as employee's company
                employer = request.user.employer_set.all().first()
                if employer:
                    employee['company'] = employer.company.id \
                        if employer.company else None
                    employee['employer'] = employer.id

                emp_serializer = EmployeeSerializer(data=employee)
                if emp_serializer.is_valid():
                    emp_serializer.save()
                    created_employees.append(emp_serializer.data)
                else:
                    raise ValidationError({
                        "message": "Invalid Employee's Arguments",
                        "total_created": len(created_employees),
                        "created_employees": created_employees
                    })
            # to validate employee's account details
            #django_rq.enqueue(background_jobs.validate_bank_account_details)
            return JsonResponse({
                    "message": "Employee List Successfully Created.",
                    "total_created": len(created_employees),
                    "created_employees": created_employees
                })
        else:
            raise ValidationError({
                "message": "Missing fields param."
            })


class EmployeeStatusView(views.APIView):
    response_schema_dict = {
        "200": openapi.Response(
            description="success",
            examples={
                "application/json": {
                    "phone": "string",
                    "registered": "true / false",
                    "confirmed": "true / false",
                    "otp": "string"
                }
            }
        ),
    }

    phone_param = openapi.Parameter('phone', openapi.IN_QUERY, description="Enter phone number", type=openapi.TYPE_STRING)

    @swagger_auto_schema(
        method='get', manual_parameters=[phone_param], responses=response_schema_dict)
    @action(detail=False, methods=['GET'])
    def get(self, request):
        phone = request.GET.get('phone')
        user = User.objects.filter(username=phone).first()
        if user:
            employee = user.employee_set.filter(deleted_at=None)
        else:
            employee = []

        # get otp and send to the user
        otp = utilities.generate_random_number(6)
        template = "login"
        data = call(phone, template, otp)

        if user and employee:
            user.otp = otp
            expiry_datetime = datetime.datetime.now() + datetime.timedelta(seconds=120)
            expiry_datetime = get_utc_datetime(expiry_datetime)
            user.otp_valid_till = expiry_datetime
            user.save()
            return JsonResponse(
                {
                    "phone": str(employee[0].phone),
                    "registered": True,
                    "confirmed": employee[0].confirmed,
                    "otp": str(otp)
                }
            )
        if user and not employee:
            return JsonResponse(
                {
                    "phone": phone,
                    "registered": False,
                    "confirmed": False,
                    "otp": str(otp)
                }
            )
        if not user:
            return JsonResponse(
                {
                    "phone": phone,
                    "registered": False,
                    "confirmed": False,
                    "otp": str(otp)
                }
            )


